import os
import math
import pandas as pd
import streamlit as st
import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv
from google.cloud import vision
import re
import base64
import cv2
import numpy as np
from PIL import Image
import io
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import qrcode
import tempfile
import arabic_reshaper
from bidi.algorithm import get_display




# إضافات لازمة للتبويب الذكي
from rapidfuzz import process, fuzz
import time
import openpyxl

# ---- الإعدادات العامة / البيئة ----
load_dotenv()

USERNAME = "admin"
PASSWORD = "Moraqip@123"

st.set_page_config(page_title="المراقب الذكي", layout="wide")

# ---- إعداد Google Vision من secrets ----
def setup_google_vision():
    try:
        key_b64 = st.secrets["GOOGLE_VISION_KEY_B64"]
        key_bytes = base64.b64decode(key_b64)
        with open("google_vision.json", "wb") as f:
            f.write(key_bytes)
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "google_vision.json"
        return vision.ImageAnnotatorClient()
    except Exception as e:
        st.error(f"❌ لم يتم تحميل مفتاح Google Vision بشكل صحيح: {e}")
        return None

# ---- اتصال قاعدة البيانات ----
def get_conn():
    return psycopg2.connect(
        dbname=os.environ.get("DB_NAME"),
        user=os.environ.get("DB_USER"),
        password=os.environ.get("DB_PASSWORD"),
        host=os.environ.get("DB_HOST"),
        port=os.environ.get("DB_PORT"),
        sslmode=os.environ.get("DB_SSLMODE", "require")
    )

# ---- دالة تحويل الجنس ----
def map_gender(x):
    try:
        val = int(float(x))
        return "F" if val == 1 else "M"
    except:
        return "M"

# ---- تسجيل الدخول ----
def login():
    st.markdown(
        """
        <style>
        .login-container {
            display: flex;
            justify-content: center;
            align-items: flex-start; /* يرفع الصندوق لفوق */
            height: 100vh;
            padding-top: 10vh;       /* مسافة من فوق */
        }
        .login-box {
            background: #ffffff;
            padding: 1.5rem 2rem;
            border-radius: 12px;
            box-shadow: 0px 2px 12px rgba(0,0,0,0.1);
            text-align: center;
            width: 300px;
        }
        .stTextInput>div>div>input {
            text-align: center;
            font-size: 14px;
            height: 35px;
        }
        .stButton button {
            background: linear-gradient(90deg, #4e73df, #1cc88a);
            color: white;
            border-radius: 6px;
            padding: 0.4rem 0.8rem;
            font-size: 14px;
            font-weight: bold;
            transition: 0.2s;
            width: 100%;
        }
        .stButton button:hover {
            background: linear-gradient(90deg, #1cc88a, #4e73df);
            transform: scale(1.02);
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown('<div class="login-container"><div class="login-box">', unsafe_allow_html=True)

    st.markdown("### 🔑 تسجيل الدخول")
    u = st.text_input("👤 اسم المستخدم", key="login_user")
    p = st.text_input("🔒 كلمة المرور", type="password", key="login_pass")

    # ✅ كبسة واحدة تكفي
    login_btn = st.button("🚀 دخول", key="login_btn")
    if login_btn:
        if u == USERNAME and p == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()   # إعادة تحميل الصفحة مباشرة
        else:
            st.error("❌ اسم المستخدم أو كلمة المرور غير صحيحة")

    st.markdown('</div></div>', unsafe_allow_html=True)


# ---- تحقق من حالة الجلسة ----
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login()
    st.stop()

# ========================== الواجهة بعد تسجيل الدخول ==========================
tab_browse, tab_single, tab_file, tab_file_name_center, tab_count, tab_check, tab_count_custom, tab_qr, tab_cards_pdf, tab_pdf_to_table = st.tabs(
    [
        "📄 تصفّح السجلات",
        "🔍 بحث برقم",
        "📂 رفع ملف Excel",
        "🔎 البحث الذكي (اسم + مركز اقتراع)",
        "📦 عدّ البطاقات",
        "🧾 التحقق من المعلومات",
        "🧮 تحليل البيانات (COUNT)",
        "🧾 توليد QR PDF",
        "🖼️ مطابقة البطاقات → PDF",
        "🧾 PDF → جدول"   # 👈 التاب الجديد
    ]
)


# ========= تحميل خط Amiri لدعم العربية ==========
FONT_DIR = "fonts"
FONT_PATH = os.path.join(FONT_DIR, "Amiri-Regular.ttf")

# إنشاء مجلد الخطوط إذا لم يكن موجود
if not os.path.exists(FONT_DIR):
    os.makedirs(FONT_DIR)

# تسجيل الخط العربي
def register_amiri():
    try:
        pdfmetrics.registerFont(TTFont("Amiri", FONT_PATH))
        return "Amiri"
    except Exception:
        return "Helvetica"  # بديل في حال الخط ناقص

arabic_font = register_amiri()

# ========= دالة لمعالجة النص العربي ==========
def fix_arabic_text(text):
    try:
        reshaped_text = arabic_reshaper.reshape(text)
        bidi_text = get_display(reshaped_text)
        return bidi_text
    except:
        return text





# ----------------------------------------------------------------------------- #
# 1) 📄 تصفّح السجلات
# ----------------------------------------------------------------------------- #
with tab_browse:
    st.subheader("📄 تصفّح السجلات مع فلاتر")

    if "page" not in st.session_state:
        st.session_state.page = 1
    if "filters" not in st.session_state:
        st.session_state.filters = {"voter": "", "name": "", "center": ""}

    colf1, colf2, colf3, colf4 = st.columns([1,1,1,1])
    with colf1:
        voter_filter = st.text_input("🔢 رقم الناخب:", value=st.session_state.filters["voter"])
    with colf2:
        name_filter = st.text_input("🧑‍💼 الاسم:", value=st.session_state.filters["name"])
    with colf3:
        center_filter = st.text_input("🏫 مركز الاقتراع:", value=st.session_state.filters["center"])
    with colf4:
        page_size = st.selectbox("عدد الصفوف", [10, 20, 50, 100], index=1)

    if st.button("🔎 تطبيق الفلاتر"):
        st.session_state.filters = {
            "voter": voter_filter.strip(),
            "name": name_filter.strip(),
            "center": center_filter.strip(),
        }
        st.session_state.page = 1

    # --- بناء شروط البحث ---
    where_clauses, params = [], []
    if st.session_state.filters["voter"]:
        where_clauses.append('CAST("رقم الناخب" AS TEXT) ILIKE %s')
        params.append(f"%{st.session_state.filters['voter']}%")
    if st.session_state.filters["name"]:
        where_clauses.append('"الاسم الثلاثي" ILIKE %s')
        params.append(f"%{st.session_state.filters['name']}%")
    if st.session_state.filters["center"]:
        where_clauses.append('"اسم مركز الاقتراع" ILIKE %s')
        params.append(f"%{st.session_state.filters['center']}%")

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    count_sql = f'SELECT COUNT(*) FROM "Bagdad" {where_sql};'
    offset = (st.session_state.page - 1) * page_size
    data_sql = f'''
        SELECT
            "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
            "اسم مركز الاقتراع","رقم مركز الاقتراع",
            "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
        FROM "Bagdad"
        {where_sql}
        ORDER BY "رقم الناخب" ASC
        LIMIT %s OFFSET %s;
    '''

    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute(count_sql, params)
            total_rows = cur.fetchone()[0]

        df = pd.read_sql_query(data_sql, conn, params=params + [page_size, offset])
        conn.close()

        if not df.empty:
            df = df.rename(columns={
                "رقم الناخب": "رقم الناخب",
                "الاسم الثلاثي": "الاسم",
                "الجنس": "الجنس",
                "هاتف": "رقم الهاتف",
                "رقم العائلة": "رقم العائلة",
                "اسم مركز الاقتراع": "مركز الاقتراع",
                "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                "المدينة": "المدينة",
                "رقم مركز التسجيل": "رقم مركز التسجيل",
                "اسم مركز التسجيل": "اسم مركز التسجيل",
                "تاريخ الميلاد": "تاريخ الميلاد"
            })
            df["الجنس"] = df["الجنس"].apply(map_gender)

        total_pages = max(1, math.ceil(total_rows / page_size))

        # ✅ عرض النتائج
        st.dataframe(df, use_container_width=True, height=500)

        c1, c2, c3 = st.columns([1,2,1])
        with c1:
            if st.button("⬅️ السابق", disabled=(st.session_state.page <= 1)):
                st.session_state.page -= 1
                st.experimental_rerun()
        with c2:
            st.markdown(f"<div style='text-align:center;font-weight:bold'>صفحة {st.session_state.page} من {total_pages}</div>", unsafe_allow_html=True)
        with c3:
            if st.button("التالي ➡️", disabled=(st.session_state.page >= total_pages)):
                st.session_state.page += 1
                st.experimental_rerun()

    except Exception as e:
        st.error(f"❌ خطأ أثناء التصفح: {e}")

# ----------------------------------------------------------------------------- #
# 2) 🔍 البحث برقم واحد
# ----------------------------------------------------------------------------- #
with tab_single:
    st.subheader("🔍 البحث برقم الناخب")
    voter_input = st.text_input("ادخل رقم الناخب:")
    if st.button("بحث"):
        try:
            conn = get_conn()
            query = """
                SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                       "اسم مركز الاقتراع","رقم مركز الاقتراع",
                       "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                FROM "Bagdad" WHERE "رقم الناخب" LIKE %s
            """
            df = pd.read_sql_query(query, conn, params=(voter_input.strip(),))
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "رقم الناخب": "رقم الناخب",
                    "الاسم الثلاثي": "الاسم",
                    "الجنس": "الجنس",
                    "هاتف": "رقم الهاتف",
                    "رقم العائلة": "رقم العائلة",
                    "اسم مركز الاقتراع": "مركز الاقتراع",
                    "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                    "المدينة": "المدينة",
                    "رقم مركز التسجيل": "رقم مركز التسجيل",
                    "اسم مركز التسجيل": "اسم مركز التسجيل",
                    "تاريخ الميلاد": "تاريخ الميلاد"
                })
                df["الجنس"] = df["الجنس"].apply(map_gender)

                st.dataframe(df, use_container_width=True, height=500)
            else:
                st.warning("⚠️ لم يتم العثور على نتائج")
        except Exception as e:
            st.error(f"❌ خطأ: {e}")

# ----------------------------------------------------------------------------- #
# 3) 📂 رفع ملف Excel (معدل مع الأرقام غير الموجودة)
# ----------------------------------------------------------------------------- #
with tab_file:
    st.subheader("📂 البحث باستخدام ملف Excel")
    uploaded_file = st.file_uploader("📤 ارفع ملف (رقم الناخب)", type=["xlsx"])
    if uploaded_file and st.button("🚀 تشغيل البحث"):
        try:
            voters_df = pd.read_excel(uploaded_file, engine="openpyxl")
            voter_col = "رقم الناخب" if "رقم الناخب" in voters_df.columns else "VoterNo"
            voters_list = voters_df[voter_col].astype(str).tolist()

            conn = get_conn()
            placeholders = ",".join(["%s"] * len(voters_list))
            query = f"""
                SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                       "اسم مركز الاقتراع","رقم مركز الاقتراع",
                       "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                FROM "Bagdad" WHERE "رقم الناخب" IN ({placeholders})
            """
            df = pd.read_sql_query(query, conn, params=voters_list)
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "رقم الناخب": "رقم الناخب",
                    "الاسم الثلاثي": "الاسم",
                    "الجنس": "الجنس",
                    "هاتف": "رقم الهاتف",
                    "رقم العائلة": "رقم العائلة",
                    "اسم مركز الاقتراع": "مركز الاقتراع",
                    "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                    "المدينة": "المدينة",
                    "رقم مركز التسجيل": "رقم مركز التسجيل",
                    "اسم مركز التسجيل": "اسم مركز التسجيل",
                    "تاريخ الميلاد": "تاريخ الميلاد"
                })
                df["الجنس"] = df["الجنس"].apply(map_gender)

                df["رقم المندوب الرئيسي"] = ""
                df["الحالة"] = 0
                df["ملاحظة"] = ""
                df["رقم المحطة"] = 1

                df = df[["رقم الناخب","الاسم","الجنس","رقم الهاتف",
                         "رقم العائلة","مركز الاقتراع","رقم مركز الاقتراع","رقم المحطة",
                         "رقم المندوب الرئيسي","الحالة","ملاحظة"]]

                # ✅ إيجاد الأرقام غير الموجودة
                found_numbers = set(df["رقم الناخب"].astype(str).tolist())
                missing_numbers = [num for num in voters_list if num not in found_numbers]

                # عرض النتائج الموجودة
                st.dataframe(df, use_container_width=True, height=500)

                # ملف النتائج الموجودة
                output_file = "نتائج_البحث.xlsx"
                df.to_excel(output_file, index=False, engine="openpyxl")
                wb = load_workbook(output_file)
                wb.active.sheet_view.rightToLeft = True
                wb.save(output_file)
                with open(output_file, "rb") as f:
                    st.download_button("⬇️ تحميل النتائج", f,
                        file_name="نتائج_البحث.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # عرض وتحميل الأرقام غير الموجودة
                if missing_numbers:
                    st.warning("⚠️ الأرقام التالية لم يتم العثور عليها في قاعدة البيانات:")
                    st.write(missing_numbers)

                    missing_df = pd.DataFrame(missing_numbers, columns=["الأرقام غير الموجودة"])
                    miss_file = "missing_numbers.xlsx"
                    missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                    with open(miss_file, "rb") as f:
                        st.download_button("⬇️ تحميل الأرقام غير الموجودة", f,
                            file_name="الأرقام_غير_الموجودة.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.warning("⚠️ لا يوجد نتائج")
        except Exception as e:
            st.error(f"❌ خطأ: {e}")
# ----------------------------------------------------------------------------- #
# 4️⃣ التبويب الرابع: البحث الذكي بالاسم + مركز الاقتراع (Bagdad)
# ----------------------------------------------------------------------------- #
import streamlit as st
import pandas as pd
import time, tempfile, openpyxl, os
from rapidfuzz import process, fuzz

# from your_module import get_conn  # تأكد من وجودها

# ======================= أدوات التطبيع =======================
def normalize_ar(text: str) -> str:
    if not text:
        return ""
    s = str(text)
    s = s.translate(str.maketrans('', '', ''.join([
        '\u0610','\u0611','\u0612','\u0613','\u0614','\u0615','\u0616','\u0617','\u0618','\u0619','\u061A',
        '\u064B','\u064C','\u064D','\u064E','\u064F','\u0650','\u0651','\u0652','\u0653','\u0654','\u0655',
        '\u0656','\u0657','\u0658','\u0659','\u065A','\u065B','\u065C','\u065D','\u065E','\u065F','\u0670'
    ])))
    s = s.replace("ـ", "").replace(" ", "").strip()
    s = (s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
         .replace("ؤ","و").replace("ئ","ي").replace("ى","ي").replace("ة","ه"))
    return s.lower()

def normalize_fast(s: pd.Series) -> pd.Series:
    uniq = s.fillna("").astype(str).unique()
    mapping = {u: normalize_ar(u) for u in uniq}
    return s.fillna("").astype(str).map(mapping)

# ======================= تحميل انتقائي من DB =======================
@st.cache_data(show_spinner=False)
def load_baghdad_for_centers(centers: list) -> pd.DataFrame:
    conn = get_conn()
    try:
        parts = []
        batch_size = 800
        for i in range(0, len(centers), batch_size):
            batch = [str(c).strip() for c in centers[i:i+batch_size] if c]
            if not batch:
                continue
            q = """
                SELECT "رقم الناخب","الاسم الثلاثي","اسم مركز الاقتراع"
                FROM "Bagdad"
                WHERE CAST("اسم مركز الاقتراع" AS TEXT) = ANY(%(centers)s)
            """
            part = pd.read_sql_query(q, conn, params={"centers": batch})
            if not part.empty:
                parts.append(part)
        if parts:
            return pd.concat(parts, ignore_index=True)
        return pd.DataFrame(columns=["رقم الناخب","الاسم الثلاثي","اسم مركز الاقتراع"])
    finally:
        conn.close()

# ======================= واجهة Streamlit =======================
st.title("🔎 مطابقة الناخبين - بغداد (سريع)")

# 🔑 أعطِ مفاتيح فريدة لتجنب التصادم مع شاشات/صفحات/أقسام أخرى
file_nc = st.file_uploader(
    "📤 ارفع ملف Excel يحتوي الاسم + اسم مركز الاقتراع",
    type=["xlsx"],
    key="baghdad_file_uploader_v1"
)
run_nc = st.button("🚀 بدء البحث ومشاهدة التقدم", key="baghdad_run_btn_v1")

MIN_NAME_MATCH = 60

if file_nc and run_nc:
    start = time.time()
    st.info("📦 جاري تجهيز البيانات...")

    # -------- قراءة الملف --------
    try:
        df = pd.read_excel(file_nc, engine="openpyxl")
        df.columns = df.columns.str.strip()
        if not {"الاسم", "اسم مركز الاقتراع"}.issubset(df.columns):
            st.error("❌ الملف يجب أن يحتوي على الأعمدة: الاسم واسم مركز الاقتراع")
            st.stop()
    except Exception as e:
        st.error(f"❌ خطأ في قراءة الملف: {e}")
        st.stop()

    total = len(df)
    if total == 0:
        st.warning("⚠️ الملف فارغ.")
        st.stop()

    # -------- تطبيع الملف --------
    df["__norm_name"] = normalize_fast(df["الاسم"])
    df["__norm_center"] = normalize_fast(df["اسم مركز الاقتراع"])

    # -------- تحميل من القاعدة للمراكز فقط --------
    centers = df["اسم مركز الاقتراع"].dropna().astype(str).unique().tolist()
    db_all = load_baghdad_for_centers(centers)
    if db_all.empty:
        st.warning("⚠️ لا توجد بيانات مطابقة لهذه المراكز في قاعدة بغداد.")
        st.stop()

    # -------- تطبيع القاعدة + فهارس --------
    db_all["__norm_name"] = normalize_fast(db_all["الاسم الثلاثي"])
    db_all["__norm_center"] = normalize_fast(db_all["اسم مركز الاقتراع"])

    groups = {}
    for c, sub in db_all.groupby("__norm_center", sort=False):
        names = sub["__norm_name"].tolist()
        data = sub[["الاسم الثلاثي", "رقم الناخب", "اسم مركز الاقتراع", "__norm_name"]].reset_index(drop=True)
        exact = {}
        for idx, nm in enumerate(names):
            exact.setdefault(nm, []).append(idx)
        groups[c] = {"names": names, "data": data, "exact": exact}

    # -------- واجهة التقدم --------
    progress = st.progress(0)
    status = st.empty()
    log_box = st.empty()
    st.info(f"🚀 بدء المطابقة... (عدد السجلات: {total})")

    cols = df.columns
    ix_name = cols.get_loc("الاسم")
    ix_center = cols.get_loc("اسم مركز الاقتراع")
    ix_norm_name = cols.get_loc("__norm_name")
    ix_norm_center = cols.get_loc("__norm_center")

    results = []

    # -------- حلقة المطابقة --------
    for i, row in enumerate(df.itertuples(index=False, name=None), start=1):
        orig_name = row[ix_name]
        orig_center = row[ix_center]
        norm_name = row[ix_norm_name]
        norm_center = row[ix_norm_center]

        match_row = {
            "الاسم في الملف": str(orig_name),
            "النسبة تطابق الاسم": 0,
            "اسم من القاعدة": "—",
            "تطابق المدرسة": "—",
            "نسبة تطابق المدرسة": 0,
            "اسم مركز الاقتراع في القاعدة": "—",
            "رقم الناخب": "—"
        }

        grp = groups.get(norm_center)
        if grp:
            exact_hits = grp["exact"].get(norm_name)
            if exact_hits:
                idx = exact_hits[0]
                rec = grp["data"].iloc[idx]
                match_row.update({
                    "النسبة تطابق الاسم": 100,
                    "اسم من القاعدة": rec["الاسم الثلاثي"],
                    "اسم مركز الاقتراع في القاعدة": rec["اسم مركز الاقتراع"],
                    "رقم الناخب": rec["رقم الناخب"],
                    "تطابق المدرسة": "✅",
                    "نسبة تطابق المدرسة": 100
                })
            else:
                scores = process.cdist(
                    [norm_name], grp["names"],
                    scorer=fuzz.token_sort_ratio,
                    workers=-1,
                    score_cutoff=MIN_NAME_MATCH
                )[0]
                if scores.size:
                    best_idx = int(scores.argmax())
                    best_score = float(scores[best_idx])
                    if best_score >= MIN_NAME_MATCH:
                        rec = grp["data"].iloc[best_idx]
                        match_row.update({
                            "النسبة تطابق الاسم": round(best_score, 2),
                            "اسم من القاعدة": rec["الاسم الثلاثي"],
                            "اسم مركز الاقتراع في القاعدة": rec["اسم مركز الاقتراع"],
                            "رقم الناخب": rec["رقم الناخب"],
                            "تطابق المدرسة": "✅",
                            "نسبة تطابق المدرسة": 100
                        })

        results.append(match_row)

        if (i % 50 == 0) or (i == total):
            progress.progress(i / total)
            safe_name = "" if pd.isna(orig_name) else str(orig_name)
            log_box.text(f"🔹 {i}/{total}: {safe_name[:25]} ...")
            status.text(f"⏱️ تمت معالجة {i}/{total} | الوقت المنقضي: {time.time() - start:.1f} ثانية")

    # -------- العرض والحفظ --------
    final_df = pd.DataFrame(results)
    st.dataframe(final_df, use_container_width=True, height=420, key="baghdad_table_v1")

    out_file = "نتائج_البحث_الذكي.xlsx"
    final_df.to_excel(out_file, index=False)

    with open(out_file, "rb") as f:
        st.download_button("⬇️ تحميل النتائج", f, file_name=out_file, key="baghdad_download_final_v1")

    st.success(f"✅ اكتمل البحث في {time.time() - start:.1f} ثانية", icon="✅")

# ----------------------------------------------------------------------------- #
# 5) 📦 عدّ البطاقات (أرقام 8 خانات) + بحث في القاعدة + قائمة الأرقام غير الموجودة
# ----------------------------------------------------------------------------- #
with tab_count:
    st.subheader("📦 عدّ البطاقات (أرقام 8 خانات) — بحث في القاعدة + الأرقام غير الموجودة")

    imgs_count = st.file_uploader(
        "📤 ارفع صور الصفحات (قد تحتوي أكثر من بطاقة)",
        type=["jpg","jpeg","png"],
        accept_multiple_files=True,
        key="ocr_count"
    )

    if imgs_count and st.button("🚀 عدّ البطاقات والبحث"):
        client = setup_google_vision()
        if client is None:
            st.error("❌ خطأ في إعداد Google Vision.")
        else:
            all_numbers, number_to_files, details = [], {}, []

            for img in imgs_count:
                try:
                    content = img.read()
                    image = vision.Image(content=content)
                    response = client.text_detection(image=image)
                    texts = response.text_annotations
                    full_text = texts[0].description if texts else ""

                    # استخراج أرقام مكونة من 8 خانات فقط
                    found_numbers = re.findall(r"\b\d{8}\b", full_text)
                    for n in found_numbers:
                        all_numbers.append(n)
                        number_to_files.setdefault(n, set()).add(img.name)

                    details.append({
                        "اسم الملف": img.name,
                        "عدد البطاقات (أرقام 8 خانات)": len(found_numbers),
                        "الأرقام المكتشفة (أرقام 8 خانات فقط)": ", ".join(found_numbers) if found_numbers else "لا يوجد"
                    })

                except Exception as e:
                    st.warning(f"⚠️ خطأ أثناء معالجة صورة {img.name}: {e}")

            total_cards = len(all_numbers)
            unique_numbers = sorted(list(set(all_numbers)))

            st.success("✅ تم الاستخراج الأولي للأرقام")

            # ----------------- بحث في قاعدة البيانات عن الأرقام الموجودة -----------------
            found_df = pd.DataFrame()
            missing_list = []
            if unique_numbers:
                try:
                    conn = get_conn()
                    placeholders = ",".join(["%s"] * len(unique_numbers))
                    query = f"""
                        SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                               "اسم مركز الاقتراع","رقم مركز الاقتراع",
                               "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                        FROM "Bagdad" WHERE "رقم الناخب" IN ({placeholders})
                    """
                    found_df = pd.read_sql_query(query, conn, params=unique_numbers)
                    conn.close()

                    if not found_df.empty:
                        found_df = found_df.rename(columns={
                            "رقم الناخب": "رقم الناخب",
                            "الاسم الثلاثي": "الاسم",
                            "الجنس": "الجنس",
                            "هاتف": "رقم الهاتف",
                            "رقم العائلة": "رقم العائلة",
                            "اسم مركز الاقتراع": "مركز الاقتراع",
                            "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                            "المدينة": "المدينة",
                            "رقم مركز التسجيل": "رقم مركز التسجيل",
                            "اسم مركز التسجيل": "اسم مركز التسجيل",
                            "تاريخ الميلاد": "تاريخ الميلاد"
                        })
                        found_df["الجنس"] = found_df["الجنس"].apply(map_gender)

                        # 🧩 إضافة نفس الأعمدة مثل تبويب 📂 رفع ملف Excel
                        found_df["رقم المندوب الرئيسي"] = ""
                        found_df["الحالة"] = 0
                        found_df["ملاحظة"] = ""
                        found_df["رقم المحطة"] = 1

                        # ترتيب الأعمدة
                        found_df = found_df[[
                            "رقم الناخب","الاسم","الجنس","رقم الهاتف",
                            "رقم العائلة","مركز الاقتراع","رقم مركز الاقتراع","رقم المحطة",
                            "رقم المندوب الرئيسي","الحالة","ملاحظة"
                        ]]

                    found_numbers_in_db = set(found_df["رقم الناخب"].astype(str).tolist()) if not found_df.empty else set()
                    for n in unique_numbers:
                        if n not in found_numbers_in_db:
                            files = sorted(list(number_to_files.get(n, [])))
                            missing_list.append({"رقم_الناخب": n, "المصدر(الصور)": ", ".join(files)})

                except Exception as e:
                    st.error(f"❌ خطأ أثناء البحث في قاعدة البيانات: {e}")
            else:
                st.info("ℹ️ لم يتم العثور على أي أرقام مكوّنة من 8 خانات في الصور المرفوعة.")

            # ----------------- عرض النتائج للمستخدم -----------------
            st.markdown("### 📊 ملخص الاستخراج")
            c1, c2, c3 = st.columns(3)
            c1.metric("إجمالي الأرقام (مع التكرار)", total_cards)
            c2.metric("إجمالي الأرقام الفريدة (8 خانات)", len(unique_numbers))
            c3.metric("عدد الصور المرفوعة", len(imgs_count))

            st.markdown("### 🔎 بيانات الناخبين (الموجودة في قاعدة البيانات)")
            if not found_df.empty:
                st.dataframe(found_df, use_container_width=True, height=400)
                out_found = "بيانات_الناخبين_الموجودين.xlsx"
                found_df.to_excel(out_found, index=False, engine="openpyxl")
                wb = load_workbook(out_found)
                wb.active.sheet_view.rightToLeft = True
                wb.save(out_found)
                with open(out_found, "rb") as f:
                    st.download_button("⬇️ تحميل بيانات الناخبين الموجودة", f,
                        file_name="بيانات_الناخبين_الموجودين.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("⚠️ لم يتم العثور على أي مطابقات في القاعدة.")

            st.markdown("### ❌ الأرقام غير الموجودة في القاعدة (مع اسم الصورة)")
            if missing_list:
                missing_df = pd.DataFrame(missing_list)
                st.dataframe(missing_df, use_container_width=True)
                miss_file = "missing_numbers_with_files.xlsx"
                missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                with open(miss_file, "rb") as f:
                    st.download_button("⬇️ تحميل الأرقام غير الموجودة مع المصدر", f,
                        file_name="الأرقام_غير_الموجودة_مع_المصدر.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.success("✅ لا توجد أرقام مفقودة (كل الأرقام الموجودة تم إيجادها في قاعدة البيانات).")

# ----------------------------------------------------------------------------- #
# 6) 🧾 التحقق من صحة المعلومات (بواسطة باسم)
# ----------------------------------------------------------------------------- #
with tab_check:
    st.subheader("🧾 التحقق من صحة بيانات الناخبين (بواسطة باسم ⚡)")

    st.markdown("""
    **📋 التعليمات:**
    - الملف يجب أن يحتوي الأعمدة التالية:
      1. رقم الناخب  
      2. الاسم  
      3. رقم العائلة  
      4. رقم مركز الاقتراع
    """)

    uploaded_check = st.file_uploader("📤 ارفع ملف Excel للتحقق", type=["xlsx"], key="check_file")

    if uploaded_check and st.button("🚀 بدء التحقق السريع بواسطة باسم"):
        try:
            df_check = pd.read_excel(uploaded_check, engine="openpyxl")

            required_cols = ["رقم الناخب", "الاسم", "رقم العائلة", "رقم مركز الاقتراع"]
            missing_cols = [c for c in required_cols if c not in df_check.columns]

            if missing_cols:
                st.error(f"❌ الملف ناقص الأعمدة التالية: {', '.join(missing_cols)}")
            else:
                # شريط التقدم
                progress_bar = st.progress(0, text="🤖 باسم يحضّر البيانات...")
                total_steps = 4

                # تحميل أرقام الناخبين
                df_check = df_check.astype(str)
                voter_list = df_check["رقم الناخب"].str.replace(r"\.0$", "", regex=True).tolist()
                progress_bar.progress(1/total_steps, text="📥 تحميل أرقام الناخبين...")

                # جلب البيانات من القاعدة
                conn = get_conn()
                placeholders = ",".join(["%s"] * len(voter_list))
                query = f"""
                    SELECT "رقم الناخب","الاسم الثلاثي","رقم العائلة","رقم مركز الاقتراع"
                    FROM "Bagdad"
                    WHERE "رقم الناخب" IN ({placeholders})
                """
                df_db = pd.read_sql_query(query, conn, params=voter_list)
                conn.close()

                df_db = df_db.astype(str)
                df_db["رقم الناخب"] = df_db["رقم الناخب"].str.replace(r"\.0$", "", regex=True)
                progress_bar.progress(2/total_steps, text="💾 تم جلب البيانات من القاعدة...")

                # الدمج
                merged = pd.merge(df_check, df_db, on="رقم الناخب", how="left",
                                  suffixes=("_المدخل", "_القاعدة"))

                # تصحيح القيم وتحويل None
                for col in merged.columns:
                    merged[col] = merged[col].astype(str).str.replace("None", "").str.replace("nan", "").str.strip()
                    merged[col] = merged[col].str.replace(r"\.0$", "", regex=True)

                progress_bar.progress(3/total_steps, text="🧠 جاري مقارنة البيانات...")

                # التحقق
                def match(a, b): return "✅" if a == b else "❌"
                merged["تطابق الاسم"] = merged.apply(lambda r: match(r["الاسم"], r["الاسم الثلاثي"]), axis=1)
                merged["تطابق رقم العائلة"] = merged.apply(lambda r: match(r["رقم العائلة_المدخل"], r["رقم العائلة_القاعدة"]), axis=1)
                merged["تطابق المركز"] = merged.apply(lambda r: match(r["رقم مركز الاقتراع_المدخل"], r["رقم مركز الاقتراع_القاعدة"]), axis=1)

                def overall(row):
                    if row["الاسم الثلاثي"] == "":
                        return "❌ غير موجود في القاعدة"
                    elif all(row[x] == "✅" for x in ["تطابق الاسم", "تطابق رقم العائلة", "تطابق المركز"]):
                        return "✅ مطابق"
                    else:
                        return "⚠️ اختلاف"

                merged["النتيجة النهائية"] = merged.apply(overall, axis=1)

                # ملخص النتائج
                total = len(merged)
                match_count = (merged["النتيجة النهائية"] == "✅ مطابق").sum()
                diff_count = (merged["النتيجة النهائية"] == "⚠️ اختلاف").sum()
                not_found = (merged["النتيجة النهائية"] == "❌ غير موجود في القاعدة").sum()

                st.info(f"""
                ### 📊 ملخص التحقق
                - ✅ عدد السجلات المطابقة: **{match_count}**
                - ⚠️ عدد السجلات التي تحتوي اختلاف: **{diff_count}**
                - ❌ غير موجود في القاعدة: **{not_found}**
                - 📄 إجمالي السجلات: **{total}**
                """)

                # عرض النتائج
                st.dataframe(merged, use_container_width=True, height=450)

                # تحميل
                out_file = "نتائج_التحقق_السريع.xlsx"
                merged.to_excel(out_file, index=False, engine="openpyxl")
                with open(out_file, "rb") as f:
                    st.download_button("⬇️ تحميل نتائج التحقق", f,
                        file_name=out_file, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                progress_bar.progress(1.0, text="✅ التحقق اكتمل!")

        except Exception as e:
            st.error(f"❌ خطأ أثناء التنفيذ: {e}")


# ----------------------------------------------------------------------------- #
# 8) 🧩 تحليل مخصص (Group & Count)
# ----------------------------------------------------------------------------- #
st.markdown("---")
st.subheader("🧩 تحليل مخصص (Group & Count)")

st.markdown("""
**📋 التعليمات:**
- ارفع ملف Excel يحتوي الأعمدة المطلوبة.  
- اختر الأعمدة التي تريد تجميع النتائج بناءً عليها (مثلاً: *رقم مركز الاقتراع + رقم مركز التسجيل*).  
- اختر العمود الذي تريد حساب عدد تكراراته (COUNT).  
- باسم سيُظهر لك عدد الصفوف (الناخبين مثلاً) ضمن كل مجموعة 👇
""")

uploaded_group = st.file_uploader("📤 ارفع ملف Excel للتحليل المخصص", type=["xlsx"], key="group_file")

if uploaded_group:
    try:
        df = pd.read_excel(uploaded_group, engine="openpyxl")
        st.success(f"✅ تم تحميل الملف ({len(df)} صف)")

        st.markdown("### 🧱 الأعمدة المتوفرة:")
        st.write(list(df.columns))

        group_cols = st.multiselect("📊 اختر الأعمدة للتجميع (Group By):", options=df.columns)
        count_col = st.selectbox("🔢 اختر العمود المراد عده (COUNT):", options=df.columns)

        if group_cols and count_col and st.button("🚀 تنفيذ التحليل المخصص"):
            progress = st.progress(0, text="🤖 باسم يحلل البيانات...")
            total_steps = 3

            # الخطوة 1️⃣ - تجهيز البيانات
            progress.progress(1/total_steps, text="🧮 تجميع البيانات...")

            # الخطوة 2️⃣ - حساب عدد الصفوف حسب الأعمدة المحددة
            grouped = df.groupby(group_cols)[count_col].count().reset_index()
            grouped = grouped.rename(columns={count_col: "عدد الصفوف"})

            progress.progress(2/total_steps, text="📊 تجهيز النتائج...")

            # الخطوة 3️⃣ - عرض وتحميل النتائج
            st.dataframe(grouped, use_container_width=True, height=450)

            # زر تحميل النتائج
            out_file = "نتائج_تحليل_مخصص.xlsx"
            grouped.to_excel(out_file, index=False, engine="openpyxl")
            with open(out_file, "rb") as f:
                st.download_button("⬇️ تحميل النتائج (Excel)", f,
                    file_name="نتائج_تحليل_مخصص.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            progress.progress(1.0, text="✅ تم التحليل بنجاح بواسطة باسم!")
    except Exception as e:
        st.error(f"❌ حدث خطأ أثناء التحليل: {e}")

# ----------------------------------------------------------------------------- #
# 9) 🧾 توليد PDF QR (24 QR في الصفحة + دعم عربي كامل + محاذاة وسط)
# ----------------------------------------------------------------------------- #


with tab_qr:
    st.subheader("🧾 توليد ملف PDF يحتوي على QR Codes")

    uploaded_qr = st.file_uploader("📤 ارفع ملف Excel فيه (الاسم - مندوب رئيسي - رمز QR)", type=["xlsx"], key="qr_file")

    if uploaded_qr:
        try:
            import pandas as pd
            df_qr = pd.read_excel(uploaded_qr, engine="openpyxl")
        except Exception as e:
            st.error(f"❌ تعذّر قراءة الملف: {e}")
            st.stop()

        # ✅ التحقق من الأعمدة
        required_cols = ["الاسم", "مندوب رئيسي", "رمز QR"]
        if not all(col in df_qr.columns for col in required_cols):
            st.error("❌ يجب أن يحتوي الملف على الأعمدة: الاسم، مندوب رئيسي، رمز QR")
            st.stop()

        # ✅ زر إنشاء PDF
        if st.button("🚀 إنشاء ملف PDF"):
            try:
                pdf_name = "qr_codes.pdf"
                c = canvas.Canvas(pdf_name, pagesize=A4)

                rows, cols = 6, 4        # لا مشكلة نبقيها 6×4 إذا صغّرنا الـQR تلقائيًا
                page_w, page_h = A4

                # هوامش
                x_margin, y_margin = 50, 70

                # مساحة النص أسفل كل QR (سطران بخط 11 pt + هامش بسيط)
                font_size = 11
                text_block_h = 2*font_size + 12   # ≈ 34pt

                # المساحة المتاحة
                avail_w = page_w - 2*x_margin
                avail_h = page_h - 2*y_margin

                # حجم الخلية (يقسم الصفحة شبكة rows×cols)
                cell_w = avail_w / cols
                cell_h = avail_h / rows

                # احسب حجم الـQR المناسب تلقائيًا بحيث لا يتجاوز الخلية بعد خصم النص
                # اترك 6pt padding حول الـQR داخل الخلية
                inner_pad = 6
                qr_size = min(cell_w - 2*inner_pad, cell_h - text_block_h - 2*inner_pad)
                qr_size = max(40, qr_size)  # حد أدنى معقول

                count = 0
                for _, row in df_qr.iterrows():
                    name = fix_arabic_text(str(row["الاسم"]))
                    rep  = fix_arabic_text(str(row["مندوب رئيسي"]))
                    link = str(row["رمز QR"])

                    if count % (rows*cols) == 0 and count != 0:
                        c.showPage()

                    qr_img = qrcode.make(link)
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    qr_img.save(tmp.name)

                    # فهارس الخلية
                    r = (count % (rows*cols)) // cols
                    cidx = (count % (rows*cols)) % cols

                    # إحداثيات ركن الخلية العلوي الأيسر
                    cell_x = x_margin + cidx * cell_w
                    cell_top = page_h - y_margin - r * cell_h

                    # موضع الـQR: نوسّطه أفقياً وعمودياً داخل الخلية، مع إبقاء النص أسفله
                    x = cell_x + (cell_w - qr_size) / 2
                    y = cell_top - text_block_h - inner_pad - qr_size  # يبدأ من أعلى الخلية نزولاً

                    c.drawImage(tmp.name, x, y, width=qr_size, height=qr_size)

                    # نصوص أسفل الـQR داخل نفس الخلية
                    c.setFont(arabic_font, font_size)
                    text_x = cell_x + cell_w/2
                    c.drawCentredString(text_x, y - 14, name)
                    c.drawCentredString(text_x, y - 28, rep)

                    count += 1

                c.save()


                with open(pdf_name, "rb") as f:
                    st.download_button("⬇️ تحميل ملف PDF", f, file_name=pdf_name, mime="application/pdf")

                st.success("✅ تم إنشاء ملف QR PDF بنجاح!")

            except Exception as e:
                st.error(f"❌ حدث خطأ أثناء إنشاء PDF: {e}")

# ----------------------------------------------------------------------------- #
# 🧾 PDF → جدول (رقم الحزب + عمودي "ت" و "ع ص" من جميع الجداول/الصفحات)
# ----------------------------------------------------------------------------- #
# يعتمد على: pdfplumber (لاستخراج الجداول النصية) + PyMuPDF/OpenCV/Google Vision (للـ OCR fallback)

import io, re, sys
import numpy as np
import pandas as pd
import cv2

# نحاول استيراد pdfplumber و PyMuPDF مع رسائل واضحة إن لم تتوفّر
_have_pdfplumber = True
_have_pymupdf = True
try:
    import pdfplumber
except Exception:
    _have_pdfplumber = False
try:
    import fitz  # PyMuPDF
except Exception:
    _have_pymupdf = False

with tab_pdf_to_table:
    st.subheader("🧾 PDF → جدول (رقم الحزب + ت/ع ص)")

    pdf_file = st.file_uploader("📤 ارفع ملف PDF", type=["pdf"], key="pdf_to_table_party")
    run_btn  = st.button("🚀 استخراج الآن", type="primary", key="pdf_to_table_run")

    # ---------------- أدوات مساعدة عامة ----------------
    def ar_clean(s: str) -> str:
        if s is None: return ""
        s = str(s)
        s = s.replace(" ", "").replace("\u200f","").replace("\u200e","")
        s = (s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
               .replace("ؤ","و").replace("ئ","ي").replace("ى","ي").replace("ة","ه"))
        return s

    def find_party_number(text: str) -> str | None:
        if not text: return None
        m = re.search(r"اسم\s*الحزب\s*[:：]\s*(\d+)", text)
        return m.group(1) if m else None

    def pick_cols_te_as(df: pd.DataFrame) -> pd.DataFrame | None:
        if df is None or df.empty:
            return None
        # تنظيف رؤوس الأعمدة
        cols = [ar_clean(c if c else f"c{i}") for i, c in enumerate(df.columns)]
        col_map = dict(zip(df.columns, cols))
        # مرشّحات
        te_candidates = [c for c, cc in col_map.items() if cc == "ت"]
        as_candidates = [c for c, cc in col_map.items() if ("ع" in cc and "ص" in cc)]
        # احتياطي عددي
        if not te_candidates:
            numeric_like = [c for c in df.columns if pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.8]
            te_candidates = numeric_like[:1] if numeric_like else []
        if not as_candidates:
            numeric_like = [c for c in df.columns if pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.8]
            if len(numeric_like) >= 2:
                as_candidates = [numeric_like[-1]]
        if not te_candidates or not as_candidates:
            return None

        te_col, as_col = te_candidates[0], as_candidates[0]
        out = df[[te_col, as_col]].copy()
        out.columns = ["ت", "ع ص"]

        for c in ["ت","ع ص"]:
            out[c] = out[c].astype(str).str.strip().replace({"": None, "nan": None, "None": None})
        out["ت"]   = pd.to_numeric(out["ت"], errors="coerce")
        out["ع ص"] = pd.to_numeric(out["ع ص"], errors="coerce")
        out = out.dropna(subset=["ت","ع ص"]).astype({"ت":"int64","ع ص":"int64"}).reset_index(drop=True)
        return out

    # ---------------- خطوط OCR مساعدة ----------------
    def party_no_from_page_pdfplumber(pdf_obj, page_index: int) -> str | None:
        try:
            txt = pdf_obj.pages[page_index].extract_text() or ""
            return find_party_number(txt)
        except Exception:
            return None

    def party_no_from_image_bytes(img_bytes: bytes) -> str | None:
        client = setup_google_vision()
        if client is None:
            return None
        image = vision.Image(content=img_bytes)
        resp = client.text_detection(image=image)
        txt = resp.text_annotations[0].description if resp.text_annotations else ""
        return find_party_number(txt)

    def render_page_png(doc, pno, dpi=230) -> bytes:
        mat = fitz.Matrix(dpi/72, dpi/72)
        pix = doc[pno-1].get_pixmap(matrix=mat, alpha=False)
        return pix.tobytes("png")

    # --- كاشف شبكة الجداول + OCR للخلايا ---
    def detect_table_cells(img_bgr):
        """
        يرجّع (rows, flat) حيث rows = قائمة صفوف، كل صف قائمة خلايا (x,y,w,h) مرتبة.
        """
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        thr = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                    cv2.THRESH_BINARY_INV, 31, 9)
        H, W = gray.shape[:2]
        vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(20, H//35)))
        hori_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(20, W//35), 1))
        v_lines = cv2.morphologyEx(thr, cv2.MORPH_OPEN, vert_kernel, iterations=1)
        h_lines = cv2.morphologyEx(thr, cv2.MORPH_OPEN, hori_kernel, iterations=1)
        grid = cv2.add(v_lines, h_lines)

        contours, _ = cv2.findContours(grid, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        boxes = []
        for cnt in contours:
            x,y,w,h = cv2.boundingRect(cnt)
            if w < 25 or h < 20: 
                continue
            if w*h < 800: 
                continue
            if w > W*0.95 or h > H*0.95: 
                continue
            boxes.append((x,y,w,h))
        if not boxes:
            return [], []

        boxes = sorted(boxes, key=lambda b: (b[1], b[0]))  # sort by y then x
        rows = []
        current = [boxes[0]]
        for b in boxes[1:]:
            if abs(b[1] - current[-1][1]) < 15:
                current.append(b)
            else:
                rows.append(sorted(current, key=lambda r: r[0]))
                current = [b]
        rows.append(sorted(current, key=lambda r: r[0]))

        max_cols = max(len(r) for r in rows)
        norm_rows = [r for r in rows if len(r) >= max(2, int(0.6*max_cols))]
        flat = [c for r in norm_rows for c in r]
        return norm_rows, flat

    def ocr_cell(bgr, box) -> str:
        x,y,w,h = box
        pad = 4
        y0 = max(0, y+pad); x0 = max(0, x+pad)
        y1 = min(bgr.shape[0], y+h-pad); x1 = min(bgr.shape[1], x+w-pad)
        crop = bgr[y0:y1, x0:x1]
        ok, buf = cv2.imencode(".png", crop)
        if not ok: 
            return ""
        client = setup_google_vision()
        if client is None:
            return ""
        image = vision.Image(content=buf.tobytes())
        resp = client.text_detection(image=image)
        return (resp.text_annotations[0].description if resp.text_annotations else "").strip()

    def normalize_header(txt):
        txt = (txt or "").strip()
        txt = txt.replace(" ", "")
        txt = txt.replace("أ","ا").replace("إ","ا").replace("آ","ا").replace("ى","ي").replace("ة","ه")
        return txt

    # ---------------- التنفيذ الرئيسي ----------------
    if run_btn and pdf_file is not None:
        data = pdf_file.read()

        # إن لم يوجد pdfplumber ولا PyMuPDF لا يمكن المتابعة
        if not _have_pdfplumber and not _have_pymupdf:
            st.error("❌ تحتاج لتثبيت المكتبتين: `pdfplumber` و `pymupdf` لتشغيل هذا التبويب.")
            st.stop()

        final_rows = []  # سنجمع النتيجة هنا

        # --------- المرحلة 1: pdfplumber (لو متاح) ----------
        used_pdfplumber = False
        if _have_pdfplumber:
            try:
                with pdfplumber.open(io.BytesIO(data)) as pdf_obj:
                    total_pages = len(pdf_obj.pages)
                    progress = st.progress(0, text="🔎 محاولة استخراج الجداول بالنص…")

                    # إن كنا سنحتاج OCR للرقم، نفتح PyMuPDF عند الضرورة
                    doc = fitz.open(stream=data, filetype="pdf") if _have_pymupdf else None

                    for pno, page in enumerate(pdf_obj.pages, start=1):
                        party_no = party_no_from_page_pdfplumber(pdf_obj, pno-1)
                        if party_no is None and _have_pymupdf:
                            try:
                                img_bytes = render_page_png(doc, pno, dpi=230)
                                party_no = party_no_from_image_bytes(img_bytes)
                            except Exception:
                                party_no = None

                        tables = page.extract_tables({
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines",
                            "intersection_tolerance": 50
                        }) or []

                        for t_idx, tb in enumerate(tables, start=1):
                            clean = [[(c or "").strip() for c in row] for row in tb if any((c or "").strip() for c in row)]
                            if not clean:
                                continue
                            header = clean[0]
                            body   = clean[1:] if len(clean) > 1 else []
                            max_len = max(len(r) for r in ([header] + body))
                            header += [""]*(max_len-len(header))
                            body   = [r + [""]*(max_len-len(r)) for r in body]
                            df_tb  = pd.DataFrame(body, columns=[h if h else f"عمود {i+1}" for i,h in enumerate(header)])

                            picked = pick_cols_te_as(df_tb)
                            if picked is None or picked.empty:
                                continue
                            picked.insert(0, "رقم الحزب", party_no if party_no is not None else "")
                            picked["#صفحة"] = pno
                            picked["#جدول_في_الصفحة"] = t_idx
                            final_rows.append(picked)

                        progress.progress(pno/total_pages, text=f"🔎 صفحة {pno}/{total_pages}…")

                used_pdfplumber = True
            except Exception as e:
                st.info(f"ℹ️ pdfplumber فشل أو الملف مصوّر: {e}")

        # إذا نجحنا واستخراجنا صفوف، نعرض وننهي
        if final_rows:
            result_df = pd.concat(final_rows, ignore_index=True)
            st.success(f"✅ تم استخراج {len(result_df)} صف بواسطة pdfplumber.")
            st.dataframe(result_df, use_container_width=True, height=420)
            out_xlsx = "نتائج_الاحزاب_والمرشحين.xlsx"
            with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="نتائج")
                ws = writer.book["نتائج"]; ws.sheet_view.rightToLeft = True
            with open(out_xlsx, "rb") as f:
                st.download_button("⬇️ تحميل Excel", f, file_name=out_xlsx,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            out_csv = "نتائج_الاحزاب_والمرشحين.csv"
            result_df.to_csv(out_csv, index=False, encoding="utf-8-sig")
            with open(out_csv, "rb") as f:
                st.download_button("⬇️ تحميل CSV", f, file_name=out_csv, mime="text/csv")
            st.stop()

        # --------- المرحلة 2: OCR GRID (fallback) ----------
        if not _have_pymupdf:
            st.error("❌ لم يتم العثور على جداول، وPyMuPDF غير مثبت، فلا يمكن تشغيل OCR Grid Parser.\nثبت: `pymupdf`")
            st.stop()

        st.info("ℹ️ سنجرب محلّل الجداول بالصور (OCR Grid Parser)…")
        doc = fitz.open(stream=data, filetype="pdf")

        ocr_rows_all = []
        total_pages = doc.page_count
        progress2 = st.progress(0, text="🔧 OCR Grid…")

        for pno in range(1, total_pages+1):
            try:
                png_bytes = render_page_png(doc, pno, dpi=260)
                img = cv2.imdecode(np.frombuffer(png_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)

                rows, flat = detect_table_cells(img)
                if not rows:
                    progress2.progress(pno/total_pages, text=f"صفحة {pno}/{total_pages}: لم يُكتشف جدول")
                    continue

                # نفترض أول صف رؤوس أعمدة
                header_cells = rows[0]
                headers = [normalize_header(ocr_cell(img, b)) for b in header_cells]

                idx_te = None; idx_as = None
                for i, h in enumerate(headers):
                    if h == "ت": idx_te = i
                    if ("ع" in h and "ص" in h): idx_as = i

                if idx_te is None or idx_as is None:
                    numeric_score = [0]*len(headers)
                    for r in rows[1: min(len(rows), 4)]:
                        for i, b in enumerate(r):
                            val = ocr_cell(img, b)
                            if re.fullmatch(r"\d+", val or ""): numeric_score[i] += 1
                    cand = sorted(range(len(headers)), key=lambda i: -numeric_score[i])[:2]
                    if len(cand) == 2:
                        idx_te, idx_as = cand[0], cand[1]

                if idx_te is None or idx_as is None:
                    progress2.progress(pno/total_pages, text=f"صفحة {pno}/{total_pages}: فشل تحديد أعمدة ت/ع ص")
                    continue

                # رقم الحزب (من نص الصفحة إن أمكن، وإلا OCR)
                party_no = ""
                if _have_pdfplumber:
                    try:
                        with pdfplumber.open(io.BytesIO(data)) as pdf_tmp:
                            party_no = party_no_from_page_pdfplumber(pdf_tmp, pno-1) or ""
                    except Exception:
                        party_no = ""
                if not party_no:
                    party_no = party_no_from_image_bytes(png_bytes) or ""

                # قراءة بقية الصفوف
                for r in rows[1:]:
                    if idx_te >= len(r) or idx_as >= len(r): 
                        continue
                    txt_te = ocr_cell(img, r[idx_te])
                    txt_as = ocr_cell(img, r[idx_as])
                    if not (re.fullmatch(r"\d+", txt_te or "") and re.fullmatch(r"\d+", txt_as or "")):
                        continue
                    ocr_rows_all.append({
                        "رقم الحزب": party_no,
                        "ت": int(txt_te),
                        "ع ص": int(txt_as),
                        "#صفحة": pno,
                        "#جدول_في_الصفحة": 1
                    })

                progress2.progress(pno/total_pages, text=f"صفحة {pno}/{total_pages}: تمت معالجة OCR Grid")
            except Exception as e:
                progress2.progress(pno/total_pages, text=f"صفحة {pno}/{total_pages}: خطأ {e}")

        if ocr_rows_all:
            result_df = pd.DataFrame(ocr_rows_all)
            st.success(f"✅ تم استخراج {len(result_df)} صف بواسطة OCR Grid Parser.")
            st.dataframe(result_df, use_container_width=True, height=420)

            out_xlsx = "نتائج_الاحزاب_والمرشحين_OCR.xlsx"
            with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="نتائج")
                ws = writer.book["نتائج"]; ws.sheet_view.rightToLeft = True
            with open(out_xlsx, "rb") as f:
                st.download_button("⬇️ تحميل Excel (OCR)", f, file_name=out_xlsx,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            out_csv = "نتائج_الاحزاب_والمرشحين_OCR.csv"
            result_df.to_csv(out_csv, index=False, encoding="utf-8-sig")
            with open(out_csv, "rb") as f:
                st.download_button("⬇️ تحميل CSV (OCR)", f, file_name=out_csv, mime="text/csv")
        else:
            st.error("❌ لم نتمكن من استخراج جداول حتى عبر OCR Grid. ارفع نسخة أوضح/بدقّة أعلى أو أرسل عيّنة لأضبط العتبات.")
